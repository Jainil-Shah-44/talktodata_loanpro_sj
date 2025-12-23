# app/api/v1/endpoints/bucket_summary.py
import json
from datetime import datetime
from uuid import uuid4, UUID

from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from typing import List, Optional

from sqlalchemy.sql.functions import current_user

# from uuid import UUID

from app.core.database import get_db
from app.models import models
from app.models.bucket_config import BucketConfig
from app.schemas.bucket_summary import BucketSummaryRequest, BucketSummaryResponse, BucketConfigItem, \
    BucketConfigCreate, BucketConfigUpdate
from app.services.bucket_summary_service import get_multiple_bucket_summaries,get_configs
from app.core.auth.dependencies import get_current_user
from app.services.record_fields_service import get_table_columns, extract_jsonb_columns, merge_columns, is_json_col

# Added by jainil
from openpyxl import Workbook
from fastapi.responses import StreamingResponse



router = APIRouter()

@router.get("/{dataset_id}/bucket-configs",response_model=List[BucketConfigItem])
def get_bucket_summaries(
        dataset_id: str,
        db: Session = Depends(get_db),
        current_user: models.User = Depends(get_current_user)):
    configs = get_configs(
        db=db,
        dataset_id=dataset_id,
        user_id=current_user.id
    )
    return configs



@router.post("/{dataset_id}/bucket-summary", response_model=List[BucketSummaryResponse])
async def bucket_summary_endpoint(
    dataset_id:str,
    payload: BucketSummaryRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    summaries = await get_multiple_bucket_summaries(
        db=db,
        config_ids=payload.config_ids,
        config_types=payload.config_types,
        filters=payload.filters,
        user_id=current_user.id,
        dataset_id=dataset_id,
        show_empty_buckets=payload.show_empty_buckets
    )

    return summaries

#Added hvb @ 02/12/2025 for crud
@router.post("/{dataset_id}/file-bucket-configs", response_model=BucketConfigItem)
def create_bucket_config(
    dataset_id: str,
    payload: BucketConfigCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if is_dataset_provided(payload.dataset_id):
        is_json = is_json_col(dataset_id, "loan_records", "additional_fields", payload.target_field)
    else:
        is_json = payload.target_field_is_json
        dataset_id = None

    cfg = BucketConfig(
        id=uuid4(),
        dataset_id=dataset_id,
        user_id=current_user.id,
        name=payload.name,
        summary_type=payload.summary_type,  # includes your “file type logic”
        target_field=payload.target_field,
        #save as string rather then json
        #bucket_config=json.dumps(payload.bucket_config),
        target_field_is_json = is_json,
        bucket_config=payload.bucket_config,
        is_default=payload.is_default,
    )
    db.add(cfg)
    db.commit()
    db.refresh(cfg)
    return cfg

# @router.post("/default-bucket-configs", response_model=BucketConfigItem)
# def create_bucket_config(
#     payload: BucketConfigCreate,
#     db: Session = Depends(get_db),
#     current_user: models.User = Depends(get_current_user)
# ):
#     dataset_id = payload.dataset_id
#     if is_dataset_provided(dataset_id):
#         is_json = is_json_col(dataset_id, "loan_records", "additional_fields", payload.target_field)
#     else:
#         is_json = payload.target_field_is_json
#         dataset_id = None
#
#     cfg = BucketConfig(
#         id=uuid4(),
#         user_id=current_user.id,
#         name=payload.name,
#         summary_type=payload.summary_type,  # includes your “file type logic”
#         target_field=payload.target_field,
#         target_field_is_json=is_json,
#         #saves as string rather then json
#         #bucket_config=json.dumps(payload.bucket_config),
#         bucket_config=payload.bucket_config,
#         is_default=payload.is_default,
#         dataset_id=dataset_id
#     )
#     db.add(cfg)
#     db.commit()
#     db.refresh(cfg)
#     return cfg
#


@router.put("/bucket-configs/{config_id}", response_model=BucketConfigItem)
def update_bucket_config(
    config_id: UUID,
    payload: BucketConfigUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    cfg = db.query(BucketConfig).filter(
        BucketConfig.id == config_id,
        BucketConfig.user_id == current_user.id
    ).first()

    if not cfg:
        raise HTTPException(404, "Bucket config not found")

    if payload.name is not None:
        cfg.name = payload.name

    if payload.target_field is not None:
        cfg.target_field = payload.target_field

    if payload.bucket_config is not None:
        #saving as string rather then jsonb
        #cfg.bucket_config = json.dumps(payload.bucket_config)
        cfg.bucket_config = payload.bucket_config

    if payload.is_default is not None:
        cfg.is_default = payload.is_default

    cfg.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(cfg)
    return cfg

@router.delete("/bucket-configs/{config_id}")
def delete_bucket_config(
    config_id: UUID,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    cfg = db.query(BucketConfig).filter(
        BucketConfig.id == config_id,
        BucketConfig.user_id == current_user.id
    ).first()

    if not cfg:
        raise HTTPException(404, "Bucket config not found")

    db.delete(cfg)
    db.commit()
    return {"message": "Deleted"}

@router.get("/{dataset_id}/check-config")
def check_config(dataset_id, summaryType: str, targetField: str, db: Session = Depends(get_db)):

    if is_dataset_provided(dataset_id):
        exists = db.query(BucketConfig).filter(
            BucketConfig.dataset_id == dataset_id,
            BucketConfig.summary_type == summaryType,
            BucketConfig.target_field == targetField,
        ).count() > 0
    else:
        exists = db.query(BucketConfig).filter(
            BucketConfig.dataset_id.is_(None),
            BucketConfig.summary_type == summaryType,
            BucketConfig.target_field == targetField,
        ).count() > 0


    return {"exists": exists}

@router.get("/lookup-config")
def lookup_config(dataset_id: Optional[str], summary_type: str, target_field: str, db: Session = Depends(get_db)):
    q = db.query(BucketConfig).filter(
        BucketConfig.summary_type == summary_type,
        BucketConfig.target_field == target_field
    )

    if is_dataset_provided(dataset_id):
        q = q.filter(BucketConfig.dataset_id == dataset_id)
    else:
        q = q.filter(BucketConfig.dataset_id.is_(None))

    found = q.first()
    if not found:
        raise HTTPException(404, "Bucket config not found")
    else:
     return {
        "id": found.id,
        "dataset_id": found.dataset_id,
        "is_default": found.is_default,
        "summary_type": found.summary_type,
        "target_field": found.target_field
    }

@router.get("/{dataset_id}/fields-list")
def field_list(dataset_id, db: Session = Depends(get_db)):

    # Extract loan-fields
    loan_fields = get_table_columns(db, "loan_records", ["id", "dataset_id"], ["dpd",
                                                                        "dpd_as_per_string", "dpd_as_on_31st_jan_2025",
                                                                        "collection_12m", "m12_collection",
                                                                        "principal_os_amt", "disbursement_amount",
                                                                        "state", "m6_collection"])
    # Extract extra's fields
    if is_dataset_provided(dataset_id):
        json_extras = extract_jsonb_columns(db,"loan_records", dataset_id, "dataset_id","additional_fields")
    else:
        json_extras = []

    # merge json & loan_recods columns
    if len(json_extras) > 0 and len(loan_fields) > 0:
        merged_fields = merge_columns(json_extras, loan_fields)
    elif len(loan_fields) > 0:
        return loan_fields
    elif len(json_extras) > 0:
        return json_extras
    else:
        return []


    return merged_fields

def is_dataset_provided(dataset_id:str)->bool:

    if not dataset_id:
        return False
    elif dataset_id == "default":
        return False
    elif dataset_id == "":
        return False
    else:
        try:
            dataset_uuid = UUID(dataset_id)
            print(f"***********Dataset ID converted to UUID: {dataset_uuid}")
            return True
        except ValueError:
            return False
        

# added by jainil @22-12-2025        

CRORE_DIVISOR = 10_000_000

def to_crore(value):
    if value is None:
        return 0
    return round(value / CRORE_DIVISOR, 2)

@router.post("/{dataset_id}/export-bucket-summaries")
async def export_bucket_summaries(
    dataset_id: str,
    payload: BucketSummaryRequest,
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    summaries = await get_multiple_bucket_summaries(
        db=db,
        config_ids=payload.config_ids,
        config_types=payload.config_types,
        filters=payload.filters,
        user_id=user.id,
        dataset_id=dataset_id,
        show_empty_buckets=payload.show_empty_buckets,
    )    

        # Business priority (adjust if needed)
    PRIORITY_ORDER = [
        "POS",
        "DPD",
        "COLLECTION",
    ]

    def summary_priority(summary):
        name = summary["name"].lower()
        for idx, key in enumerate(PRIORITY_ORDER):
            if key.lower() in name:
                return idx
        return len(PRIORITY_ORDER)  # everything else last

    summaries.sort(key=summary_priority)


    print("==== RAW SUMMARY ORDER ====")
    for s in summaries:
        print("SUMMARY:", s["name"], "ID:", s["id"])



    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Bucket Summary"

    bold = Font(bold=True)

    row_idx = 1  # Excel rows are 1-based

    for summary in summaries:
        # ─────────────────────────────
        # Bucket Group Name (SECTION HEADER)
        # ─────────────────────────────
        ws.cell(row=row_idx, column=1, value=summary["name"]).font = bold
        row_idx += 1

        # Column headers
        headers = [
            "Bucket Label",
            "Count",
            "POS (₹ Cr)",
            "POS %",
            "Post NPA Collection (₹ Cr)",
            "Post W-Off Collection (₹ Cr)",
            "M6 Collection (₹ Cr)",
            "M12 Collection (₹ Cr)",
            "Total Collection (₹ Cr)",
        ]

        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=header).font = bold

        row_idx += 1

        # ─────────────────────────────
        # Data rows (STRICT ORDER)
        # ─────────────────────────────
        for bucket in summary["buckets"]:
            ws.cell(row=row_idx, column=1, value=bucket["label"])
            ws.cell(row=row_idx, column=2, value=int(bucket["count"]))

            ws.cell(row=row_idx, column=3, value=to_crore(bucket.get("POS")))
            ws.cell(row=row_idx, column=4, value=(bucket.get("POS_Per", 0) / 100))  # numeric %
            ws.cell(row=row_idx, column=5, value=to_crore(bucket.get("Post_NPA_Coll")))
            ws.cell(row=row_idx, column=6, value=to_crore(bucket.get("Post_W_Off_Coll")))
            ws.cell(row=row_idx, column=7, value=to_crore(bucket.get("M6_Collection")))
            ws.cell(row=row_idx, column=8, value=to_crore(bucket.get("M12_Collection")))
            ws.cell(row=row_idx, column=9, value=to_crore(bucket.get("total_collection")))

            row_idx += 1

        row_idx += 2  # spacing between bucket groups

    # ─────────────────────────────
    # Excel formatting
    # ─────────────────────────────

    # POS % formatting
    for r in range(1, row_idx):
        ws.cell(row=r, column=4).number_format = "0.00%"

    # Auto-size columns
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # ----------------------------
    # Number formatting (2 decimals)
    # ----------------------------

    for r in range(1, row_idx):
        # POS % column (percentage)
        ws.cell(row=r, column=4).number_format = "0.00%"

        # Monetary columns (₹ Cr) → 2 decimals
        for c in [3, 5, 6, 7, 8, 9]:
            ws.cell(row=r, column=c).number_format = "#,##0.00"
    

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=all_bucket_summaries.xlsx"
        },
    )
