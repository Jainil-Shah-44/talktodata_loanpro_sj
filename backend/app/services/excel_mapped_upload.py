import pandas as pd
import io
from datetime import datetime
from sqlalchemy import create_engine, text
from uuid import UUID
from openpyxl import load_workbook
import json

# ============================================================
# 1️⃣ FUNCTION: READ & MAP EXCEL FILE USING CONFIG
# ============================================================
def upload_to_postgres(data_id: UUID, df, db_engine, table_name, column_mapping=None, truncate_before_insert=False):
    """
    Upload a DataFrame to PostgreSQL with robust handling of mapped and extra columns.

    Args:
        data_id: Unique identifier for this dataset upload.
        df: pandas DataFrame containing data to upload.
        db_engine: SQLAlchemy engine or connection.
        table_name: Target PostgreSQL table.
        column_mapping: Dict {df_col_name: db_col_name} for mapped columns.
        truncate_before_insert: Bool, whether to clear table before inserting.

    Returns:
        Dict: {"status": True/False, "inserted": count, "total": total_records}
    """
    try:
        if df is None or df.empty:
            return {"status": False, "inserted": 0, "total": 0, "message": "No data to upload"}

        df = df.copy()

        # Add data_id column
        df['data_id'] = str(data_id)

        # ===== Handle column mapping =====
        # this supports only 1-1 column mapping,
        # commented hvb @ 05/11/2025 replaced with 1-N column mappings
        # if column_mapping:
        #     mapped_cols = list(column_mapping.keys())
        #
        #     # Only keep columns that actually exist in df
        #     existing_mapped_cols = [c for c in mapped_cols if c in df.columns]
        #
        #     # Prepare df with mapped columns
        #     df_to_upload = df[existing_mapped_cols].copy()
        #
        #     # Rename to database column names
        #     df_to_upload = df_to_upload.rename(columns={c: column_mapping[c] for c in existing_mapped_cols})
        #
        #     # Handle extra columns as JSON
        #     # extra_cols = [c for c in df.columns if c not in existing_mapped_cols]
        #     # if extra_cols:
        #     #     # Convert datetime / other non-serializable to str
        #     #     # df_to_upload['extra_data_json'] = df[extra_cols].apply(
        #     #     #     lambda row: json.dumps({k: (v.isoformat() if hasattr(v, "isoformat") else v)
        #     #     #                             for k,v in row.items()}, ensure_ascii=False),
        #     #     #     axis=1
        #     #     # )
        #     #     df_to_upload['extra_data_json'] = df[extra_cols].apply(
        #     #         lambda row: json.dumps({k: safe_serialize(v) for k, v in row.items()}, ensure_ascii=False),
        #     #         axis=1
        #     #     )
        # else:
        #     df_to_upload = df.copy()
        # ===== Handle column mapping (supports 1→many) =====
        if column_mapping:

            # Start with an empty df that keeps the same index as original df
            df_to_upload = df[[]].copy()  # empty, but same index

            for src_col, target in column_mapping.items():
                if src_col not in df.columns:
                    continue

                # 1 → many mapping (duplicate source column to multiple DB columns)
                if isinstance(target, list):
                    for db_col in target:
                        df_to_upload[db_col] = df[src_col]
                else:
                    # Normal 1 → 1 mapping
                    df_to_upload[target] = df[src_col]

        else:
            df_to_upload = df.copy()

        # ===== Truncate table if requested =====
        if truncate_before_insert:
            with db_engine.begin() as conn:
                conn.execute(text(f"TRUNCATE TABLE {table_name} RESTART IDENTITY CASCADE"))

        # ===== Count records before insert =====
        with db_engine.begin() as conn:
            result = conn.execute(text(f"SELECT COUNT(*) FROM {table_name}"))
            total_before = result.scalar()

        # ===== Upload =====
        df_to_upload.to_sql(table_name, con=db_engine, if_exists="append", index=False)

        # ===== Count records after insert =====
        with db_engine.begin() as conn:
            result = conn.execute(text(f"SELECT COUNT(*) FROM {table_name}"))
            total_after = result.scalar()

        inserted = total_after - total_before
        return {"status": True, "inserted": inserted, "total": len(df_to_upload)}

    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"status": False, "inserted": 0, "total": len(df) if df is not None else 0, "message": str(e)}


# ============================================================
# 2️⃣ FUNCTION:
# # =============UPLOAD DATAFRAME TO POSTGRES===============================================
# def upload_to_postgres(data_id:UUID,df, db_engine, table_name, column_mapping=None, truncate_before_insert=False,create_new_table=False):
#     """
#     Uploads a DataFrame to PostgreSQL using SQLAlchemy engine.
#
#     Args:
#         df: DataFrame containing data to upload.
#         db_engine: SQLAlchemy engine instance.
#         table_name: Target PostgreSQL table name.
#         column_mapping: Optional dict {json_col: db_col}.
#         truncate_before_insert: If True, clears table before inserting.
#
#     Returns:
#         True if successful, False otherwise.
#     """
#     try:
#         if df is None or df.empty:
#             print("⚠️ No data to upload")
#             return False
#
#         df = df.copy()
#
#         # Add data_id column
#         df['data_id'] = data_id
#
#         # --- Convert dicts (like extra_data_json) to JSON strings ---
#         for col in df.columns:
#             if df[col].apply(lambda x: isinstance(x, (dict, list))).any():
#                 df[col] = df[col].apply(
#                     lambda x: json.dumps(
#                         x,
#                         ensure_ascii=False,
#                         default=str  # 👈 This converts datetime/date/Decimal to string automatically
#                     ) if isinstance(x, (dict, list)) else x
#                 )
#
#         # Apply column mapping
#         # if column_mapping:
#         #     df = df.rename(columns=column_mapping)
#
#         if column_mapping:
#             mapped_cols = list(column_mapping.keys())
#             df_to_upload = df[mapped_cols].copy()  # keep only mapped columns
#             df_to_upload = df_to_upload.rename(columns=column_mapping)
#
#             # Optional: handle extras as JSON,un-comment following if you wish to upload un-mapped columns as extras
#             # extra_cols = [c for c in df.columns if c not in mapped_cols]
#             # if extra_cols:
#             #    df_to_upload['extra_data_json'] = df[extra_cols].apply(lambda x: x.to_dict(), axis=1)
#         else:
#             df_to_upload = df.copy()  # no mapping, upload all
#
#         # Replace NaN/None with NULL-safe values
#         # df = df.where(pd.notnull(df), None)
#         df_to_upload = df_to_upload.where(pd.notnull(df), None)
#
#         # Truncate table if requested
#         if truncate_before_insert:
#             with db_engine.begin() as conn:
#                 conn.execute(text(f"TRUNCATE TABLE {table_name} RESTART IDENTITY CASCADE"))
#                 print(f"🧹 Truncated table '{table_name}' before insert")
#
#         # --- Record count before insert ---
#         if create_new_table:
#             pre_count = 0
#         else:
#             with db_engine.begin() as conn:
#                 pre_count = conn.execute(text(f"SELECT COUNT(*) FROM {table_name}")).scalar()
#
#         # Upload data
#         # df.to_sql(table_name, con=db_engine, if_exists="append", index=False)
#         df_to_upload.to_sql(table_name, con=db_engine, if_exists="append", index=False)
#         print(f"✅ Uploaded {len(df)} records to table '{table_name}'")
#
#         # --- Record count after insert ---
#         with db_engine.begin() as conn:
#             post_count = conn.execute(text(f"SELECT COUNT(*) FROM {table_name}")).scalar()
#
#         inserted = post_count - pre_count
#         print(f"✅ Inserted {inserted} new records into '{table_name}' (Total now: {post_count})")
#
#         # return True
#
#         return {"status": True, "inserted": inserted, "total": len(df)}
#
#
#
#     except Exception as e:
#         print(f"❌ Error uploading to Postgres: {e}")
#         import traceback
#         traceback.print_exc()
#         #return False
#         return {"status": False, "inserted": 0, "total": len(df)}
# ============================================================
# 2️⃣ FUNCTION:
# # =============UPLOAD DATAFRAME TO POSTGRES===============================================
def fn_read_excel_map_base(excel_bytes, mapping_config):
    """
    Reads and maps multiple Excel sheets using flexible mapping rules.

    Args:
        excel_bytes: File content in bytes (from UploadFile.read()).
        mapping_config: Dict defining sheets, header, cols, and relations., where sheet numbers are 1-based mapping and not 0-based.
        {
            "sheets": {
                1: {
                    "header_row": 2,                                  # zero-based index for column header row, pass this as -1 for header-less excel
                                                                      # For header-less, excel column names are renamed as _alias_col_idx
                    "skip_rows": 1,                                   # rows to skip before data, this is after header in case data starts after 2 rows of header you pass this as 2
                    "cols_to_read": [0, 1, 2],                        # columns to read (by index - zero-based index), or pass "all" to read all columns
                    "datetime_headers": [1],                          # column indices with datetime headers
                    "alias": "Pool"                                   # logical name
                    "extra": [{"3": "dated"}, {"5": "amt_in_debt"}] # specify extra column to be included reading 3rd,5th column and naming them dated,amt_in_debt respectively
                    "key_columns":[0]                                 # column index for your join key(s), row will be dropped if this column contains Nulls
                },
                2: {
                    "header_row": 0,
                    "skip_rows": 0,
                    "cols_to_read": "all",
                    "alias": "DPD"
                },
                3: {
                    "header_row": 1,
                    "skip_rows": 0,
                    "cols_to_read": [0, 2, 5],
                    "alias": "Collection"
                }
            },
            "relations": [
                {"left": 1, "right": 2, "left_col": 0, "right_col": 3, "how": "left"},
                {"left": 1, "right": 3, "left_col": 0, "right_col": 1, "how": "left"}
            ]
        }
    Returns:
        DataFrame with joined and cleaned data, ready for upload.
    """
    # mod hvb @ 23/11/2025 removed exception handler let callee function handle it.
    #try:
    start_time = datetime.now()
    print(f"📘 Starting Excel read at {start_time}")

    excel_file = pd.ExcelFile(io.BytesIO(excel_bytes))
    dfs = {}


    suffix_sheet_num_to_extra = True
    if mapping_config["sheets"]:
        if len(mapping_config["sheets"]) == 1:
            suffix_sheet_num_to_extra = False

    # === STEP 1: Read each sheet as per mapping ===
    for sheet_no, cfg in mapping_config["sheets"].items():
        sheet_name = excel_file.sheet_names[sheet_no - 1]
        print(f"➡️ Reading sheet {sheet_no}: {sheet_name}")

        sheet_alias = cfg.get("alias", f"Sheet{sheet_no}")
        print(f"➡️ Reading sheet alias {sheet_alias}")

        # Define columns to read
        usecols = None if cfg.get("cols_to_read", "all") == "all" else cfg["cols_to_read"]

        rename_col_by_idx = False
        header_row_idx = cfg.get("header_row", -1)
        if header_row_idx < 0:
            header_row_idx = None
            rename_col_by_idx = True

        # Read Excel
        print(f"➡️ Pulling data from excel file {sheet_name} @ {datetime.now()}")
        read_start_time = datetime.now()

        # Replaced with better approach to read, reading only content
        # df = pd.read_excel(
        #     excel_file,
        #     sheet_name=sheet_name,
        #     #header=cfg.get("header_row", 0),
        #     header=header_row_idx,
        #     skiprows=cfg.get("skip_rows", 0),
        #     usecols=usecols,
        # )

        df = read_excel_data_only(excel_bytes, sheet_name, header_row_idx, cfg.get("skip_rows", 0), usecols=usecols)

        #shifting key_cols_read after setting cols_to_read
        #hvb @ 23/11/2025
        # key_column_indices = cfg.get("key_columns", [])  # define in your mapping config per sheet
        # if key_column_indices:
        #     key_cols = [df.columns[idx] for idx in key_column_indices]
        #     df = df.dropna(subset=key_cols).reset_index(drop=True)

        print(f"✅ Completed Reading sheet data {sheet_name} @ {datetime.now()}, Total time = {datetime.now() - read_start_time}")

        # Clean column headers
        if rename_col_by_idx:
            # Rename columns to _col_{index}
            # Creates indexing based on dataframe columns rather than actual columns
            # df.columns = [f"_{sheet_alias}_col_{i}" for i in range(len(df.columns))]
            if usecols is not None:
                cols_to_read = usecols
                excel_col_indices = cols_to_read
                if isinstance(cols_to_read, str):
                    if cols_to_read == "all":
                        excel_col_indices = list(range(len(df.columns)))
                    else:
                        cols_to_read = [int(c.strip()) for c in cols_to_read.split(",") if c.strip().isdigit()]
                        excel_col_indices = cols_to_read

                df.columns = [f"_{sheet_alias}_col_{i}" for i in excel_col_indices]
        else:
            df.columns = [
                str(c).strip().lower().replace(" ", "_").replace("unnamed:", "")
                for c in df.columns
            ]

        key_column_indices = cfg.get("key_columns", [])  # define in your mapping config per sheet
        if key_column_indices:
            if cols_to_read:
                key_cols = [df.columns[cols_to_read.index(idx)] for idx in key_column_indices]
            else:
                key_cols = [df.columns[idx] for idx in key_column_indices]
            df = df.dropna(subset=key_cols).reset_index(drop=True)

        # Format datetime headers if defined
        for col_index in cfg.get("datetime_headers", []):
            if col_index < len(df.columns):
                new_name = datetime.now().strftime("%Y%m%d_%H%M%S")
                df.rename(columns={df.columns[col_index]: new_name}, inplace=True)

        # Replace NaN / empty / None
        df = df.replace({pd.NA: None, "NaN": None, "nan": None, "": None})

        # Check for datetime column from config and replace them with null for non-parsed
        clean_columns = cfg.get("clean_columns", [])
        for item in clean_columns:
            for col_index, ctype in item.items():
                # Added hvb @ 23/11/2025 to get clean-up idx as we are only reading specified columns so count
                # is reduced now!
                if cols_to_read:
                    clean_idx = cols_to_read.index(col_index)
                else:
                    clean_idx = col_index
                # if col_index < len(df.columns):
                #     if ctype == "dt":
                #         clean_invalid_dates(df, [df.columns[col_index]])
                #     elif ctype == "int":
                #         clean_invalid_int(df,df.columns[col_index])
                if clean_idx < len(df.columns):
                    if ctype == "dt":
                        clean_invalid_dates(df, [df.columns[clean_idx]])
                    elif ctype == "int":
                        clean_invalid_int(df,df.columns[clean_idx])
                    elif ctype == "float":
                        clean_invalid_float(df,df.columns[clean_idx])
                    elif ctype == "string":
                        clean_invalid_string(df, df.columns[clean_idx])

        # --- Handle extras (if provided) ---
        extras_config = cfg.get("extra", None)
        if extras_config:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)
            ws = wb[sheet_name]

            # Collect all rows from sheet (for extra cols only)
            raw_data = list(ws.iter_rows(values_only=True))
            extra_data_list = []

            for row_idx in range(len(df)):
                extra_fields = {}
                for item in extras_config:
                    for col_idx_str, alias_name in item.items():
                        col_idx = int(col_idx_str)
                        # Add only if within bounds
                        if col_idx < len(
                                raw_data[row_idx + cfg.get("skip_rows", 0) + (cfg.get("header_row", 0) + 1)]):
                            val = raw_data[row_idx + cfg.get("skip_rows", 0) + (cfg.get("header_row", 0) + 1)][
                                col_idx]
                            extra_fields[alias_name] = val
                extra_data_list.append(extra_fields)

            # Attach collected JSON as a new column
            if suffix_sheet_num_to_extra:
                df["extra_data_json_" + str(sheet_no)] = extra_data_list
            else:
                df["extra_data_json"] = extra_data_list

        #dfs[cfg.get("alias", f"Sheet{sheet_no}")] = df
        dfs[sheet_alias] = df


    # === STEP 2: Join sheets using relations ===
    combined_df = None
    for rel in mapping_config.get("relations", []):
        left_alias = mapping_config["sheets"][rel["left"]].get("alias", f"Sheet{rel['left']}")
        right_alias = mapping_config["sheets"][rel["right"]].get("alias", f"Sheet{rel['right']}")

        left_df = dfs[left_alias]
        right_df = dfs[right_alias]
        # mod hvb @ 20/11/2025 read mapping from db, this is string
        try:
            left_col_id = int(rel["left_col"])
        except ValueError:
            raise ValueError(f"Invalid left col in join : {rel['left_col']}")

        try:
            right_col_id = int(rel["right_col"])
        except ValueError:
            raise ValueError(f"Invalid right col in join : {rel['right_col']}")

        # left_col_name = left_df.columns[rel["left_col"]]
        # right_col_name = right_df.columns[rel["right_col"]]

        left_col_name = left_df.columns[left_col_id]
        right_col_name = right_df.columns[right_col_id]

        print(f"🔗 Joining {left_alias}.{left_col_name} -> {right_alias}.{right_col_name}")

        if combined_df is None or combined_df is left_df:
            combined_df = left_df.merge(
                right_df,
                how=rel.get("how", "left"),
                left_on=left_col_name,
                right_on=right_col_name
            )
        else:
            combined_df = combined_df.merge(
                right_df,
                how=rel.get("how", "left"),
                left_on=left_col_name,
                right_on=right_col_name
            )

    #Added hvb @ 23/11/2025 if combine_df is null, check size of dfs_sheet_wise and set first one as combined_df
    if combined_df is None:
        if len(dfs) == 0:
            raise ValueError("No data found")
        else:
            first_df_key = next(iter(dfs))
            first_df_value = dfs[first_df_key]
            combined_df = first_df_value

    # --- Merge extra_data_json from multiple sheets into one column ---
    extra_cols = [c for c in combined_df.columns if c.startswith("extra_data_json")]
    if extra_cols:
        def merge_extra_json(row):
            merged = {}
            for col in extra_cols:
                val = row.get(col, {})
                if isinstance(val, dict):
                    merged.update(val)
            return merged

        # merge columns
        if(len(extra_cols) > 1): #added hvb @ 24/11/2025 to merge if column count of extra is more than 1.
            combined_df["extra_data_json"] = combined_df.apply(merge_extra_json, axis=1)

        #convert data to jsonb format
        combined_df["extra_data_json"] = combined_df["extra_data_json"].apply(
            lambda v: json.dumps(v, ensure_ascii=False, default=str) if isinstance(v, dict) else v
        )

        # Optional: drop the original per-sheet extra_data_json columns
        if suffix_sheet_num_to_extra: # Added hvb @ 24/11/2025 drop suffixed column which are now merged as one.
            combined_df.drop(columns=extra_cols, inplace=True, errors="ignore")
    print(f"✅ Completed Excel read and join in {datetime.now() - start_time}")
    return combined_df
    # mod hvb @ 23/11/2025 removed exception handler let callee function handle it.
    #except Exception as e:
    #    print(f"❌ Error processing Excel: {e}")
    #    import traceback
    #    traceback.print_exc()
    #    return None
#v2

# ============================================================
# 3️⃣ EXAMPLE: END-TO-END DEMO
# ============================================================
# if __name__ == "__main__":
#     # --- Example mapping for Excel ---
#     mapping_config = {
#         "sheets": {
#             1: {"header_row": 1, "skip_rows": 0, "cols_to_read": "all", "alias": "Pool"},
#             2: {"header_row": 0, "skip_rows": 0, "cols_to_read": "all", "alias": "DPD"},
#             3: {"header_row": 1, "skip_rows": 0, "cols_to_read": "all", "alias": "Collection"}
#         },
#         "relations": [
#             {"left": 1, "right": 2, "left_col": 0, "right_col": 3, "how": "left"},
#             {"left": 1, "right": 3, "left_col": 0, "right_col": 1, "how": "left"}
#         ]
#     }
#
#     # --- Example column mapping (for upload) ---
#     column_mapping = {
#         "col_agt": "agent_id",
#         "col_zagt": "dpd_agent_id",
#         "col_ya": "collection_agent_id"
#     }
#
#     # --- Example DB Engine ---
#     # (replace credentials with your actual DB)
#     db_engine = create_engine(
#         "postgresql+psycopg2://username:password@localhost:5432/your_database"
#     )
#
#     # --- Read Excel file ---
#     with open("sample_data.xlsx", "rb") as f:
#         excel_bytes = f.read()
#
#     merged_df = fn_read_excel_map_base(excel_bytes, mapping_config)
#
#     if merged_df is not None:
#         # --- Upload merged data to Postgres ---
#         upload_to_postgres(
#             merged_df,
#             db_engine=db_engine,
#             table_name="loan_data",
#             column_mapping=column_mapping,
#             truncate_before_insert=True
#         )
#
#         print("🎯 Data successfully processed and uploaded.")
#     else:
#         print("❌ Failed to read Excel file.")

# ============================================================
# 4️⃣ Read excel as data only, avoiding reading styles, for fast load
# ============================================================

def read_excel_data_only(file_content, sheet_name, header=None, skiprows=0, usecols=None,max_empty_rows=20):
    # Load workbook in read-only mode, data_only=True ensures only cell values are read
    wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    ws = wb[sheet_name]

    # Extract pure data (no styles)
    data = []
    empty_row_counter = 0

    for row in ws.iter_rows(values_only=True):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            empty_row_counter += 1
            if empty_row_counter >= max_empty_rows:
                break  # stop reading if continuous empty rows counter has reached beyond specified limit
            continue
        else:
            empty_row_counter = 0  # reset counter on first non-empty row
            data.append(row)

    # Convert to DataFrame
    df = pd.DataFrame(data)

    # Apply header logic manually
    if header is not None:
        df.columns = df.iloc[header].tolist()
        df = df.drop(index=range(0, header + 1))

    if usecols is not None:
        cols_to_read = usecols
        if isinstance(cols_to_read, str):
            cols_to_read = [int(c.strip()) for c in cols_to_read.split(",") if c.strip().isdigit()]
        df = df.iloc[:, cols_to_read]

    # Remove any remaining fully empty rows (safety)
    df = df.dropna(how='all').reset_index(drop=True)

    if skiprows:
        df = df.iloc[skiprows:]

    df.reset_index(drop=True, inplace=True)
    return df

def safe_serialize(val):
    if val is None or pd.isna(val):
        return None
    if hasattr(val, "isoformat"):  # datetime, Timestamp
        return val.isoformat()
    if isinstance(val, (int, float, str, bool)):
        return val
    return str(val)  # fallback for other objects

from datetime import datetime

def clean_invalid_dates(df, datetime_columns):
    """
    Replaces any non-date value in specified datetime columns with None.

    Args:
        df: pandas DataFrame
        datetime_columns: list of column names (strings)

    Returns:
        Cleaned DataFrame (modifies in-place)
    """
    for col in datetime_columns:
        if col in df.columns:
            def safe_parse_date(val):
                if pd.isna(val) or val is None:
                    return None
                if isinstance(val, (datetime, pd.Timestamp)):
                    return val
                try:
                    # Try parsing common date formats
                    return pd.to_datetime(val, errors='raise')
                except Exception:
                    return None  # Invalid date string (like "Pending", "NA", etc.)

            df[col] = df[col].apply(safe_parse_date)
    return df

def clean_invalid_string(df,column):
    """
            Cleans invalid or non-string values in the given column of a DataFrame.

            - Converts valid strings to string
            - Replaces invalid / NaN values with None
            - Ensures safe values for PostgreSQL insertion
            """
    def safe_str(val):
        try:
            # Handle NaN, NaT, None etc.
            if pd.isna(val):
                return None
            # Try converting to integer (works for numeric strings like "123")
            return str(val)
        except (ValueError, TypeError):
            return None

    df[column] = df[column].apply(safe_str)
    return df

def clean_invalid_float(df,column):
    """
        Cleans invalid or non-float values in the given column of a DataFrame.

        - Converts valid float strings to float
        - Replaces invalid / non-numeric / NaN values with None
        - Ensures safe values for PostgreSQL insertion
        """
    def safe_float(val):
        try:
            # Handle NaN, NaT, None etc.
            if pd.isna(val):
                return None
            # Try converting to integer (works for numeric strings like "123")
            return float(val)
        except (ValueError, TypeError):
            return None

    df[column] = df[column].apply(safe_float)
    return df

def clean_invalid_int(df, column):
    """
    Cleans invalid or non-integer values in the given column of a DataFrame.

    - Converts valid integer strings to int
    - Replaces invalid / non-numeric / NaN values with None
    - Ensures safe values for PostgreSQL insertion
    """
    def safe_int(val):
        try:
            # Handle NaN, NaT, None etc.
            if pd.isna(val):
                return None
            # Try converting to integer (works for numeric strings like "123")
            return int(val)
        except (ValueError, TypeError):
            return None

    df[column] = df[column].apply(safe_int)
    return df

